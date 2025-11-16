import os
import re
import time
import openpyxl
from urllib.parse import urljoin
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
import chromedriver_autoinstaller
import subprocess

# -------------------
# ✅ 파일명 정리 함수
# -------------------
def sanitize_filename(name):
    return re.sub(r'[\\/*?:"<>|]', "_", name)

# -------------------
# ✅ 플래그 설정
# -------------------
new_updates = False
flag_path = "./result_files/new_updates.flag"

# -------------------
# ✅ 크롬드라이버 설정
# -------------------
chromedriver_autoinstaller.install()
chrome_options = Options()
chrome_options.add_argument('--headless')
chrome_options.add_argument('--no-sandbox')
chrome_options.add_argument('--disable-dev-shm-usage')
driver = webdriver.Chrome(service=Service(), options=chrome_options)
driver.implicitly_wait(5)

# -------------------
# ✅ 저장 경로 설정
# -------------------
save_base = './result_files'
os.makedirs(save_base, exist_ok=True)
excel_path = os.path.join(save_base, "metadata.xlsx")

# -------------------
# ✅ 엑셀 초기화 or 로드
# -------------------
if not os.path.exists(excel_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "게시글 목록"
    ws.append(["게시글 제목", "관련부서", "작성일", "URL"])
    wb.save(excel_path)
    existing_keys = set()
else:
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    existing_keys = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        title, dept, date, url = row
        if url:
            existing_keys.add(url.strip())
    print(f"✅ 기존 게시글 {len(existing_keys)}건 로드 완료.")

# -------------------
# ✅ 엑셀에 데이터 추가
# -------------------
def append_to_excel(title, dept, date, url, excel_path):
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    last_row = 1
    for row in range(ws.max_row, 0, -1):
        if any(cell.value for cell in ws[row]):
            last_row = row
            break
    ws.cell(row=last_row + 1, column=1, value=title)
    ws.cell(row=last_row + 1, column=2, value=dept)
    ws.cell(row=last_row + 1, column=3, value=date)
    ws.cell(row=last_row + 1, column=4, value=url)
    wb.save(excel_path)

# -------------------
# ✅ 크롤링 시작
# -------------------
base_url = "https://swknu.kongju.ac.kr"
board_url = f"{base_url}/community/notice.do?&pn=1"
max_pages = 1

print(f"\n========== 🔍 공주대 SW중심대학 공지사항 크롤링 시작 ==========")
driver.get(board_url)
page_num = 1

while True:
    print(f"\n📄 {page_num}페이지 처리 중... ({driver.current_url})")
    items = driver.find_elements(By.CSS_SELECTOR, ".list-photo .item")
    if not items:
        print("❌ 게시글 항목을 찾을 수 없음, 종료.")
        break

    for idx, item in enumerate(items, start=1):
        try:
            link_elem = item.find_element(By.CSS_SELECTOR, "a[href*='noticedetail.do']")
            title = item.find_element(By.CSS_SELECTOR, ".title").text.strip()
            post_url = urljoin(base_url, link_elem.get_attribute("href"))
            info_elems = item.find_elements(By.CSS_SELECTOR, ".post-info span")
            dept = "공주대학교SW중심대학사업단"
            date = info_elems[1].text.strip() if len(info_elems) > 1 else "정보 없음"

            if post_url in existing_keys:
                print(f"⏩ ({idx}) {title} → 이미 존재, 건너뜀")
                continue

            print(f"📰 ({idx}) {title} ({date}) → 새 게시글 처리 중...")
            new_updates = True

            driver.get(post_url)
            time.sleep(1)

            try:
                content_elem = driver.find_element(By.CSS_SELECTOR, ".view-note")
                content = content_elem.text.strip()
            except:
                content = "본문을 가져올 수 없습니다."

            file_links = []
            try:
                file_elems = driver.find_elements(By.CSS_SELECTOR, "div.post-file ul li a")
                for f in file_elems:
                    fname = f.text.strip()
                    furl = urljoin(base_url, f.get_attribute("href"))
                    if fname and fname != "미리보기":
                        file_links.append((fname, furl))
            except:
                pass

            safe_title = sanitize_filename(title)[:80]
            file_path = os.path.join(save_base, f"{safe_title}.md")
            markdown = f"""# {title}

**관련부서:** {dept}  
**작성일:** {date}  
**URL:** {post_url}  

---

## 본문
{content}

---

## 첨부파일
"""
            if file_links:
                for fname, furl in file_links:
                    markdown += f"- [{fname}]({furl})\n"
            else:
                markdown += "첨부파일 없음\n"

            with open(file_path, "w", encoding="utf-8") as f:
                f.write(markdown)

            append_to_excel(title, dept, date, post_url, excel_path)
            existing_keys.add(post_url)
            print(f"✅ 저장 완료 → {file_path}")

            driver.back()
            time.sleep(1)

        except Exception as e:
            print(f"❗ 게시글 처리 실패: {e}")
            continue

    try:
        next_page = driver.find_element(By.CSS_SELECTOR, f"a[href*='pn={page_num+1}']")
        driver.get(next_page.get_attribute("href"))
        page_num += 1
        if page_num > max_pages:
            print(f"🔒 최대 {max_pages}페이지 도달 → 종료")
            break
    except:
        print("📄 다음 페이지 없음 → 종료")
        break

driver.quit()
print("\n✅ 모든 크롤링 완료!")

# ✅ GitHub 자동 푸시
subprocess.run(["git", "config", "--global", "user.email", "github-actions@github.com"])
subprocess.run(["git", "config", "--global", "user.name", "github-actions"])
subprocess.run(["git", "add", "."])
subprocess.run(["git", "commit", "-m", "Auto update crawl results"])
subprocess.run(["git", "push"])

# -------------------
# 🔔 새 게시글 있으면 flag 생성
# -------------------
if new_updates:
    with open(flag_path, "w") as f:
        f.write("new")
    print("📌 새로운 게시글 존재 → 임베딩 실행 플래그 생성")
else:
    print("📌 새로운 게시글 없음 → 임베딩 스킵 예정")
